import os
from datetime import datetime
from typing import List

import openpyxl
import pandas as pd
import requests
from core.config import RECOMMENDATION_FILE_PATH
from repository.database import get_db_connection
from utils.sql_reader import read_sql_query


def upload_container(history_id, file_path):
    """
    Загружает файл на сервер по указанному пути.
    :param history_id: ID истории для формирования имени файла.
    :param file_path: Полный путь к файлу для загрузки.
    :return: Код статуса HTTP-ответа.
    """
    url = "http://webspd.mlc.gov/SpdRemote/spdremotemvc/MassStart/Parse"
    login, password = "GabitovDS", "21v1DF43!!!!"
    print('Начало загрузки файла')

    try:
        with open(file_path, "rb") as file:
            # Данные формы
            data = {
                "selectedMassObjectType": "1",
                "reglamentId": "934",
                "klassIndex": "801-402",
            }
            print(f'Файл для загрузки: {file_path}')

            # Файл для загрузки
            files = {
                "file": (f"container_{history_id}.xlsx", file, 
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            }
            print(f'Файл подготовлен: {files}')

            # Выполнение POST-запроса
            response = requests.post(url, data=data, files=files, auth=(login, password))

            print(f'Ответ сервера: {response.status_code}')
            return response.status_code
    except FileNotFoundError:
        print(f"Файл {file_path} не найден.")
        return 404
    except Exception as e:
        print(f"Произошла ошибка: {e}")
        return 500

def generate_excel_from_two_dataframes(history_id=None, output_dir="./uploads", affair_ids=None):
    """
    Генерирует Excel-файл и загружает его на сервер.
    :param history_id: ID истории для формирования имени файла.
    :param output_dir: Директория для сохранения файла.
    :param new_selected_addresses: Список новых адресов.
    :param old_selected_addresses: Список старых адресов.
    :param affair_ids: Список affair_id квартир для генерации контейнера.
    :param date: Флаг для включения даты.
    :return: Код статуса HTTP-ответа от сервера.
    """
    # Проверка и создание директории, если её нет
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Формируем полный путь к файлу
    if history_id:
        file_name = f"container_{history_id}.xlsx"
    else:
        file_name = f"container_0.xlsx"
    output_path = os.path.join(output_dir, file_name)

    # Создаем соединение с базой данных
    connection = get_db_connection()
    cursor = connection.cursor()

    # Основной запрос
    query = """
        WITH unnst AS (
            SELECT 
                offer_id,
                affair_id, 
                (KEY)::bigint as new_apart_id
            FROM offer,
            jsonb_each(new_aparts)
            WHERE offer_id IN (SELECT max(offer_id) FROM offer GROUP BY affair_id)
            ORDER BY offer_id DESC
        )
        SELECT 
            oa.full_house_address, 
            oa.apart_number, 
            oa.type_of_settlement, 
            oa.kpu_number, 
            o.new_apart_id, 
            na.house_address, 
            na.apart_number, 
            na.full_living_area, 
            na.total_living_area, 
            na.room_count, 
            na.living_area, 
            na.floor, 
            cin_address, 
            cin_schedule, 
            dep_schedule, 
            phone_osmotr, 
            phone_otvet, 
            na.entrance_number,
            (start_dates_by_entrence->>(na.entrance_number::text))::date AS start_date,
            c.otdel,
			c.full_house_address,
			c.full_cin_address,
			oa.district,
			na.cad_num,
			mail_index
        FROM unnst o
        JOIN old_apart oa USING (affair_id)
        JOIN new_apart na USING (new_apart_id) 
        JOIN test_cin c ON c.house_address = na.house_address
        JOIN mail_index on oa.full_house_address = mail_index.house_address
        WHERE oa.is_queue <> 1
    """
    params = []
        
    if history_id:
        query += " AND na.history_id = %s AND oa.history_id = %s"
        params.extend([history_id, history_id])

    # Добавляем условие по affair_ids, если они переданы
    if affair_ids:
        query += " AND oa.affair_id IN %s"
        params.append(tuple(affair_ids))

    # Выполняем запрос
    cursor.execute(query, params)
    aparts = cursor.fetchall()
    print('aparts', aparts)

    # Создаем DataFrame из результатов запроса
    df = pd.DataFrame(aparts, columns=[
        'full_old_house_address', 'old_number', 'type_of_settlement', 'kpu_number', 'new_apart_id', 
        'new_address', 'new_number', 'full_living_area', 'total_living_area', 'room_count', 
        'living_area', 'floor', 'cin_address', 'cin_schedule', 'dep_schedule', 'phone_osmotr', 'phone_otvet', 'entrance_number', 'start_date', 'otdel',
        'full_house_address', 'full_cin_address', 'old_district', 'old_cad_num', 'mail_index'
    ])

    print(df['full_living_area'], df['total_living_area'], df['living_area'])
    print(df)

    # Создаем новую книгу Excel и выбираем активный лист
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Определяем заголовки на русском и английском
    russian_headers = [
        "Номер заявки", "Заявитель.Тип (1=ФЛ,2=ЮЛ,3=ИП)", "Заявитель.Порядковый номер", "Заявитель.Имя",
        "Заявитель.Наименование юр.лица", "Заявитель.Почтовый адрес. Адрес строкой", "Заявитель.Почтовый адрес. Индекс",
        "Заявитель.Почтовый адрес. Населенный пункт", "Помещение. Кадастровый номер", "Помещение. Адресный ориентир",
        "Помещение. Площадь общая", "Помещение. Площадь жилая", "Помещение. Этаж", "Помещение. Количество комнат",
        "Помещение. Площадь Жилого помещения", "Помещение. Статус помещения", "Помещение. Идентификатор источника",
        "Свед.о действ.отнош. Номер", "Доп.параметр.Имя", "Доп.параметр.Значение", "Идентификатор документа для автогенерации"
    ]
    english_headers = [
        "NumPP", "appApplicantList.type", "appApplicantList.tab", "appApplicantList.firstname", "appApplicantList.name",
        "appApplicantList.postAddressList.address", "appApplicantList.postAddressList.index", "appApplicantList.postAddressList.locality",
        "flatList.cadastralNumber", "flatList.address", "flatList.square", "flatList.livingSquare", "flatList.floorNumber",
        "flatList.room_number", "flatList.sDwellingArea", "flatList.flatStatus", "flatList.sourceid", "actDocList.documentNumber",
        "customInputDataList.tagName", "customInputDataList.value", "reportID"
    ]

    # Заполняем первую строку русскими заголовками
    for col_num, header in enumerate(russian_headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Заполняем вторую строку английскими идентификаторами
    for col_num, header in enumerate(english_headers, 1):
        sheet.cell(row=2, column=col_num, value=header)

    # Дополнительные значения и тексты для столбцов
    additional_values = ["VSOOTVET", "INFO_SOB", "ISPOLNITEL", "OSMOTR", "GETKEY", "OTVET"]
    # Заполняем данные из DataFrame
    row_num = 3  # Начинаем с третьей строки

    for index, row in df.iterrows():
        if row['old_district'] in ("ЗелАО", "ВАО", "ЮВАО", "САО", "СВАО"):
            report_id = 108404 if row['type_of_settlement'] == "частная собственность" else 108407
        elif row['old_district'] in ("ЗАО", "СЗАО", "ЮАО", "ЮЗАО", "ТАО", "НАО"):
            report_id = 108406 if row['type_of_settlement'] == "частная собственность" else 108409
        elif row['old_district'] in ("ЦАО"):
            report_id = 108405 if row['type_of_settlement'] == "частная собственность" else 108408
        for i, tag in enumerate(additional_values):
            additional_texts = {
                "VSOOTVET": f"Согласно постановлению Правительства Москвы от 01.08.2017 № 497-ПП «О программе реновации жилищного фонда в городе Москве» (далее - Программа реновации) в отношении многоквартирного дома по адресу: г. Москва, {row['full_old_house_address']} принято решение о включении в Программу реновации.",
                "INFO_SOB": """- заявление о включении в предмет Договора предлагаемого жилого помещения;\n- оригиналы документов личного характера и правоустанавливающие документы на освобождаемое жилое помещение.\nПросим довести указанную в письме информацию до всех правообладателей.""",
                "ISPOLNITEL": "Кандабаров Н.А.",
                "OSMOTR": f"""Для осмотра квартиры необходимо {('с ' + (str(row['start_date'])[8:] + '.' + str(row['start_date'])[5:7] + '.' +str(row['start_date'])[:4])) if row['start_date'] != None and row['start_date'] > datetime.now().date() else '' } в течение 7 рабочих дней обратиться в информационный центр по адресу: г. Москва, {row['full_cin_address']} ({row['cin_schedule']}) по предварительной записи онлайн на сайте https://www.mos.ru/ в разделе «Осмотр квартиры» (для перехода наведите камеру смартфона на QR~код) или по тел. {row['phone_osmotr']}.""" if row['cin_schedule'] != 'time2plan' else "Предварительная запись на показ жилого помещения доступна на сервисе онлайн-записи «Время планировать вместе с ДГИ»: https://time2plan.online.",
                "GETKEY": " ",
                "OTVET": f"Всем правообладателям необходимо предоставить свое согласие либо отказ от предлагаемого жилого помещения в срок не позднее 7 рабочих дней в информационный центр по адресу: г. Москва, {row['full_cin_address'] if row['dep_schedule']!= 'отдел' else row['otdel']} {('(' + row['dep_schedule'] + ')') if row['dep_schedule']!= 'отдел' else ''}, по предварительной записи по вышеуказанному тел., при себе необходимо иметь следующие документы:"
            }

            # Номер заявки
            sheet.cell(row=row_num, column=1, value=index + 1)


            if i == 0:  # Первая строка каждой группы
                sheet.cell(row=row_num, column=2, value=1)  # appApplicantList.type
                sheet.cell(row=row_num, column=3, value=1)  # appApplicantList.tab
                sheet.cell(row=row_num, column=4, value="Уважаемый правообладатель!")  # appApplicantList.firstname
                sheet.cell(row=row_num, column=6, value=f"{row['full_old_house_address']} кв. {row['old_number']}")  # Адрес
                sheet.cell(row=row_num, column=7, value=row['mail_index'])  # Индекс
                sheet.cell(row=row_num, column=8, value="г. Москва")  # Населенный пункт
                sheet.cell(row=row_num, column=9, value=row['old_cad_num'])  # Кадастровый номер
                sheet.cell(row=row_num, column=10, value=f"{row['full_house_address']}, кв. {row['new_number']}")  # flatList.address

                # Заполняем данные flatList
                sheet.cell(row=row_num, column=11, value=row['total_living_area'])  # flatList.square
                sheet.cell(row=row_num, column=12, value=row['living_area'])  # flatList.livingSquare
                sheet.cell(row=row_num, column=13, value=row['floor'])  # flatList.floorNumber
                sheet.cell(row=row_num, column=14, value=row['room_count'])  # flatList.room_number
                sheet.cell(row=row_num, column=15, value=row['full_living_area'])  # flatList.sDwellingArea
                sheet.cell(row=row_num, column=16, value=3)  # flatList.flatStatus
                sheet.cell(row=row_num, column=17, value=row['new_apart_id'])  # flatList.sourceid
                sheet.cell(row=row_num, column=18, value=row['kpu_number'])  # actDocList.documentNumber
                sheet.cell(row=row_num, column=21, value=report_id)  # reportID

            sheet.cell(row=row_num, column=19, value=tag)  # customInputDataList.tagName
            sheet.cell(row=row_num, column=20, value=additional_texts[tag])  # customInputDataList.value

            row_num += 1

    # Сохраняем файл
    wb.save(output_path)
    print(f"Excel файл '{output_path}' создан.")
    connection.close()
    print('-------------------------------------------------------------\nDONE\n---------------------------------------------------------------------')

def set_is_uploaded(history_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute('UPDATE public.history set is_downloaded = True where history_id = %s', (history_id, ))
    connection.commit()
    print('DONE', history_id)
    connection.close()

def update_apart_status_by_history_id(history_id: int) -> str:
    if not isinstance(history_id, int) or history_id <= 0:
        raise ValueError("history_id должен быть положительным целым числом")
    connection = None  # Инициализируем переменную заранее
    sql_file_path = os.path.join(RECOMMENDATION_FILE_PATH, 'UpdateOfferStatusByHistoryId.sql')
    
    try:
        with get_db_connection() as connection:
            with connection.cursor() as cursor:
                # Читаем SQL из файла
                with open(sql_file_path, 'r', encoding='utf-8') as f:
                    sql_query = f.read()
                
                # Выполняем запрос в транзакции
                cursor.execute(sql_query, (history_id,))
                connection.commit()

                updated_rows = cursor.fetchone()[0]
                
                return f"Successfully updated. Affected rows: {updated_rows}"
                
    except FileNotFoundError:
        raise Exception(f"SQL файл не найден: {sql_file_path}")
    except Exception as e:
        print(e)
        if connection is not None:
            connection.rollback()
        raise Exception(f"Database error occurred: {str(e)}")
    

def update_apart_status(apart_ids: List[int]) -> str:
    if not apart_ids or not all(isinstance(i, int) for i in apart_ids):
        raise ValueError("apart_ids должен быть непустым списком целых чисел")

    connection = None
    sql_file_path = os.path.join(RECOMMENDATION_FILE_PATH, 'UpdateOfferStatusByAffairIds.sql')
    
    try:
        with get_db_connection() as connection:
            with connection.cursor() as cursor:
                with open(sql_file_path, 'r', encoding='utf-8') as f:
                    sql_query = f.read()
                
                cursor.execute(sql_query, (apart_ids,))
                connection.commit()

                result, affected_rows = cursor.fetchone()
                
                return f"Successfully updated. Result: {result}, Affected rows: {affected_rows}"
                
    except FileNotFoundError:
        raise Exception(f"SQL файл не найден: {sql_file_path}")
    except Exception as e:
        if connection is not None:
            connection.rollback()
        raise Exception(f"Database error occurred: {str(e)}")