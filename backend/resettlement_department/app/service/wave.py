import psycopg2
import pandas as pd
from core.config import settings
from datetime import datetime
import json
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from pathlib import Path

def get_db_connection():
    return psycopg2.connect(
        host=settings.project_management_setting.DB_HOST,
        user=settings.project_management_setting.DB_USER,
        password=settings.project_management_setting.DB_PASSWORD,
        port=settings.project_management_setting.DB_PORT,
        database=settings.project_management_setting.DB_NAME
       
    )

def wave_matching(
    df_new_apart,
    df_old_apart,
    cursor,
    conn
):
    try:
        max_rank_by_room_count = df_old_apart.groupby("room_count")["rank"].max().to_dict()
        df_new_apart_rev = df_new_apart.loc[::-1]
        offers_to_insert = []
        cannot_offer_to_insert = []

        old_apart_ranks = df_old_apart.groupby("room_count")["rank"].max().to_dict()
        old_apart_count = df_old_apart.groupby("room_count")["affair_id"].count().to_dict()
        new_apart_count = df_new_apart.groupby("room_count")["new_apart_id"].count().to_dict()

        df_old_apart_reversed = df_old_apart.loc[::-1]
        a = {}
        delta = {1: 1.5, 2: 3, 3: 5, 4: 6.5, 5: 8, 6: 9.5, 7: 11, 8: 12.5}

        print("df_old_apart['room_count'].max()", df_old_apart['room_count'].max(), df_new_apart['room_count'].max())

        for i in range(1, (df_old_apart['room_count'].max() if df_old_apart['room_count'].max() > df_new_apart['room_count'].max() else df_new_apart['room_count'].max()) + 1):
            if (((old_apart_ranks[i] if old_apart_ranks.get(i) is not None else  0) > (max_rank_by_room_count[i] if max_rank_by_room_count.get(i) is not None else  0)) 
            or ((old_apart_count[i] if old_apart_count.get(i) is not None else  0) > (new_apart_count[i] if new_apart_count.get(i) is not None else  0))):
                a[i] = 0

                old_apart_list = []

                print("DEFICIT")
                for _, old_apart in df_old_apart_reversed[df_old_apart_reversed["room_count"] == i].iterrows():
                    old_apart_id = int(old_apart["affair_id"])
                    # Определяем условие по этажу
                    floor_condition = (
                        ((df_new_apart["floor"] >= (old_apart["min_floor"] - 2)) & 
                            (df_new_apart["floor"] <= (old_apart["max_floor"] + 2))
                        ) if old_apart["min_floor"] or old_apart["max_floor"] else True
                    )

                
                    if (
                        old_apart["buying_date"] is not None and old_apart["buying_date"] > datetime.strptime("2017-08-01", "%Y-%m-%d").date()
                        and not (old_apart["min_floor"] or old_apart["max_floor"])
                    ):
                        s = df_new_apart[
                            (df_new_apart["room_count"]== old_apart["room_count"])&
                            (df_new_apart["full_living_area"]>= old_apart["full_living_area"])&
                            (df_new_apart["total_living_area"]>= old_apart["total_living_area"])&
                            (df_new_apart["living_area"]>= old_apart["living_area"])&
                            (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"])
                        ]

                        if not s.empty:
                            sap = s.iloc[0]

                            suitable_aparts = df_new_apart[
                                (df_new_apart["room_count"]== old_apart["room_count"])& 
                                (df_new_apart["full_living_area"]>= old_apart["full_living_area"])&
                                (df_new_apart["total_living_area"]>= old_apart["total_living_area"])& 
                                (df_new_apart["living_area"]>= old_apart["living_area"])&
                                (df_new_apart["for_special_needs_marker"]== old_apart["is_special_needs_marker"])&
                                ((df_new_apart["full_living_area"] - sap["full_living_area"])<= delta[i])
                            ].sort_values(by="floor", ascending=True)

                            # Если подходящих квартир нет, проверяем с условием floor_condition
                            if suitable_aparts.empty:
                                suitable_aparts = df_new_apart[
                                    (df_new_apart["room_count"]== old_apart["room_count"])& 
                                    (df_new_apart["full_living_area"]>= old_apart["full_living_area"])&
                                    (df_new_apart["total_living_area"]>= old_apart["total_living_area"])&
                                    (df_new_apart["living_area"]>= old_apart["living_area"])&
                                    (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"])
                                ]

                                if suitable_aparts.empty:
                                    cannot_offer_to_insert.append((old_apart_id,))
                                else:
                                    suitable_apart = suitable_aparts.iloc[0]
                                    new_apart_id = int(suitable_apart["new_apart_id"])
                                    df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]
                                    offers_to_insert.append((old_apart_id, new_apart_id))
                            else:
                                suitable_apart = suitable_aparts.iloc[0]
                                new_apart_id = int(suitable_apart["new_apart_id"])
                                df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]
                                offers_to_insert.append((old_apart_id, new_apart_id,))
                        else:
                            cannot_offer_to_insert.append((old_apart_id, ))

                    else:
                        # Если "Дата покупки" <= 2017-08-01 или пустая, используем floor_condition
                        suitable_aparts = df_new_apart[
                            (df_new_apart["room_count"]== old_apart["room_count"])&
                            (df_new_apart["full_living_area"]>= old_apart["full_living_area"])&
                            (df_new_apart["total_living_area"] >= old_apart["total_living_area"])&
                            (df_new_apart["living_area"]>= old_apart["living_area"])&
                            (df_new_apart["for_special_needs_marker"]== old_apart["is_special_needs_marker"])&
                            (df_new_apart["floor"]>= (old_apart["min_floor"]))& 
                            (df_new_apart["floor"] <= (old_apart["max_floor"]))
                        ]
                        # Проверка наличия подходящих квартир
                        if suitable_aparts.empty:
                            suitable_aparts = df_new_apart[
                                (df_new_apart["room_count"] == old_apart["room_count"])&
                                (df_new_apart["full_living_area"] >= old_apart["full_living_area"])&
                                (df_new_apart["total_living_area"] >= old_apart["total_living_area"])&
                                (df_new_apart["living_area"]>= old_apart["living_area"])&
                                (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"])&
                                floor_condition
                            ]

                            if suitable_aparts.empty:
                                suitable_aparts = df_new_apart[
                                    (df_new_apart["room_count"] == old_apart["room_count"])&
                                    (df_new_apart["full_living_area"] >= old_apart["full_living_area"])& 
                                    (df_new_apart["total_living_area"] >= old_apart["total_living_area"])&
                                    (df_new_apart["living_area"] >= old_apart["living_area"])& 
                                    (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"])
                                ]
                                if suitable_aparts.empty:
                                    cannot_offer_to_insert.append((old_apart_id, ))
                                else:
                                    suitable_apart = suitable_aparts.iloc[0]
                                    new_apart_id = int(suitable_apart["new_apart_id"])
                                    old_apart_list.append(old_apart_id)
                                    df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]
                            else:
                                suitable_apart = suitable_aparts.iloc[0]
                                new_apart_id = int(suitable_apart["new_apart_id"])
                                old_apart_list.append(old_apart_id)
                                df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]
                        else:
                            suitable_apart = suitable_aparts.iloc[0]
                            new_apart_id = int(suitable_apart["new_apart_id"])
                            old_apart_list.append(old_apart_id)
                            df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]

                df_new_apart_second = df_new_apart_rev.loc[::-1]
                print('REVERSED NEW', df_new_apart_second[df_new_apart_second['room_count'] == 2])
                print('REVERSED OLD', df_old_apart_reversed[df_old_apart_reversed["room_count"] == 2])

                for _, old_apart in df_old_apart_reversed[df_old_apart_reversed["room_count"] == i].iterrows():
                    old_apart_id = int(old_apart["affair_id"])
                    if old_apart_id not in old_apart_list:
                        continue
                    # Определяем условие по этажу
                    floor_condition = (
                        ((df_new_apart_second["floor"]>= (old_apart["min_floor"] - 2))&
                        (df_new_apart_second["floor"] <= (old_apart["max_floor"] + 2)))
                        if old_apart["min_floor"] or old_apart["max_floor"]
                        else True
                    )

                    # Условие для даты "Дата покупки"
                    if (old_apart["buying_date"] is not None and old_apart["buying_date"] > datetime.strptime("2017-08-01", "%Y-%m-%d").date()
                        and not (old_apart["min_floor"] or old_apart["max_floor"])
                    ):
                        s = df_new_apart_second[
                            (df_new_apart_second["room_count"]== old_apart["room_count"])&
                            (df_new_apart_second["full_living_area"]>= old_apart["full_living_area"])& 
                            (df_new_apart_second["total_living_area"]>= old_apart["total_living_area"])& 
                            (df_new_apart_second["living_area"]>= old_apart["living_area"])& 
                            (df_new_apart_second["for_special_needs_marker"]== old_apart["is_special_needs_marker"])
                        ]

                        if not s.empty:
                            sap = s.iloc[0]

                            suitable_aparts = df_new_apart_second[
                                (df_new_apart_second["room_count"] == old_apart["room_count"])& 
                                (df_new_apart_second["full_living_area"]>= old_apart["full_living_area"])& 
                                (df_new_apart_second["total_living_area"]>= old_apart["total_living_area"])&
                                (df_new_apart_second["living_area"]>= old_apart["living_area"])& 
                                (df_new_apart_second["for_special_needs_marker"]== old_apart["is_special_needs_marker"])& 
                                ((df_new_apart_second["full_living_area"]- sap["full_living_area"] ) <= delta[i])
                            ].sort_values(by="floor", ascending=True)

                            # Если подходящих квартир нет, проверяем с условием floor_condition
                            if suitable_aparts.empty:
                                suitable_aparts = df_new_apart_second[
                                    (df_new_apart_second["room_count"]  == old_apart["room_count"])& 
                                    (df_new_apart_second["full_living_area"]>= old_apart["full_living_area"])& 
                                    (df_new_apart_second["total_living_area"]>= old_apart["total_living_area"])& 
                                    (df_new_apart_second["living_area"] >= old_apart["living_area"])& 
                                    (df_new_apart_second[ "for_special_needs_marker"]== old_apart["is_special_needs_marker"])
                                ]

                                if suitable_aparts.empty:
                                    cannot_offer_to_insert.append((old_apart_id,))
                                else:
                                    suitable_apart = suitable_aparts.iloc[0]
                                    new_apart_id = int(suitable_apart["new_apart_id"])
                                    df_new_apart_second = df_new_apart_second[df_new_apart_second["new_apart_id"]!= new_apart_id]
                                    offers_to_insert.append(
                                        (old_apart_id, new_apart_id,)
                                    )
                            else:
                                suitable_apart = suitable_aparts.iloc[0]
                                new_apart_id = int(
                                    suitable_apart["new_apart_id"]
                                )
                                df_new_apart_second = df_new_apart_second[
                                    df_new_apart_second["new_apart_id"]
                                    != new_apart_id
                                ]
                                offers_to_insert.append((old_apart_id, new_apart_id,))
                        else:
                            cannot_offer_to_insert.append((old_apart_id, ))
                    else:
                        # Если "Дата покупки" <= 2017-08-01 или пустая, используем floor_condition
                        suitable_aparts = df_new_apart_second[
                            (df_new_apart_second["room_count"]  == old_apart["room_count"])& 
                            (df_new_apart_second["full_living_area"] >= old_apart["full_living_area"])& 
                            (df_new_apart_second["total_living_area"]>= old_apart["total_living_area"])& 
                            (df_new_apart_second["living_area"]>= old_apart["living_area"])& 
                            ( df_new_apart_second["for_special_needs_marker"] == old_apart["is_special_needs_marker"])& 
                            (df_new_apart_second["floor"] >= (old_apart["min_floor"]))& 
                            (df_new_apart_second["floor"]<= (old_apart["max_floor"]))
                        ]

                        # Проверка наличия подходящих квартир
                        if suitable_aparts.empty:
                            suitable_aparts = df_new_apart_second[
                                (df_new_apart_second["room_count"] == old_apart["room_count"])& 
                                (df_new_apart_second["full_living_area"]>= old_apart["full_living_area"])& 
                                (df_new_apart_second["total_living_area"]>= old_apart["total_living_area"])& 
                                (df_new_apart_second["living_area"]>= old_apart["living_area"])& 
                                (df_new_apart_second["for_special_needs_marker"]== old_apart["is_special_needs_marker"])
                                & floor_condition
                            ]
                            if suitable_aparts.empty:
                                suitable_aparts = df_new_apart_second[
                                    (df_new_apart_second["room_count"] == old_apart["room_count"])& 
                                    (df_new_apart_second["full_living_area"] >= old_apart["full_living_area"])& 
                                    (df_new_apart_second["total_living_area"] >= old_apart["total_living_area"])&
                                    (df_new_apart_second["living_area"]>= old_apart["living_area"])& 
                                    (df_new_apart_second["for_special_needs_marker"] == old_apart["is_special_needs_marker"])
                                ]
                                if suitable_aparts.empty:
                                    print('PROBLEM --- ',old_apart_id)
                                    cannot_offer_to_insert.append((old_apart_id,))

                                else:
                                    suitable_apart = suitable_aparts.iloc[0]
                                    new_apart_id = int(suitable_apart["new_apart_id"])
                                    df_new_apart_second = df_new_apart_second[df_new_apart_second["new_apart_id"]!= new_apart_id]
                                    offers_to_insert.append((old_apart_id, new_apart_id,))

                            else:
                                suitable_apart = suitable_aparts.iloc[0]
                                new_apart_id = int(suitable_apart["new_apart_id"])
                                df_new_apart_second = df_new_apart_second[df_new_apart_second["new_apart_id"] != new_apart_id]
                                offers_to_insert.append((old_apart_id, new_apart_id, ))
                        else:
                            suitable_apart = suitable_aparts.iloc[0]
                            new_apart_id = int(suitable_apart["new_apart_id"])
                            df_new_apart_second = df_new_apart_second[
                                df_new_apart_second["new_apart_id"] != new_apart_id]
                            offers_to_insert.append((old_apart_id, new_apart_id,))
            else:
                a[i] = 1
                print("PROFICIT")

                for _, old_apart in df_old_apart[
                    df_old_apart["room_count"] == i
                ].iterrows():
                    old_apart_id = int(old_apart["affair_id"])
                    # Определяем условие по этажу
                    floor_condition = ((
                            (df_new_apart["floor"]>= (old_apart["min_floor"] - 2))& 
                            (df_new_apart["floor"] <= (old_apart["max_floor"] + 2)))
                        if old_apart["min_floor"] or old_apart["max_floor"]
                        else True
                    )

                    # Условие для даты "Дата покупки"
                    if (old_apart["buying_date"] is not None and old_apart["buying_date"] > datetime.strptime("2017-08-01", "%Y-%m-%d").date() 
                        and not (old_apart["min_floor"] or old_apart["max_floor"])
                    ):
                        s = df_new_apart[
                            (df_new_apart["room_count"] == old_apart["room_count"]) &
                            (df_new_apart["full_living_area"] >= old_apart["full_living_area"]) &
                            (df_new_apart["total_living_area"] >= old_apart["total_living_area"]) &
                            (df_new_apart["living_area"] >= old_apart["living_area"]) &
                            (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"])
                        ]

                        if not s.empty:
                            sap = s.iloc[0]

                            suitable_aparts = df_new_apart[
                                (df_new_apart["room_count"] == old_apart["room_count"]) &
                                (df_new_apart["full_living_area"] >= old_apart["full_living_area"]) &
                                (df_new_apart["total_living_area"] >= old_apart["total_living_area"]) &
                                (df_new_apart["living_area"] >= old_apart["living_area"]) &
                                (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"]) &
                                ((df_new_apart["full_living_area"] - sap["full_living_area"]) <= delta[i])
                            ].sort_values(by="floor", ascending=True)

                            # Если подходящих квартир нет, проверяем с условием floor_condition
                            if suitable_aparts.empty:
                                suitable_aparts = df_new_apart[
                                    (df_new_apart["room_count"] == old_apart["room_count"])  &
                                    (df_new_apart["full_living_area"] >= old_apart["full_living_area"])  &
                                    (df_new_apart["total_living_area"] >= old_apart["total_living_area"]) &
                                    (df_new_apart["living_area"] >= old_apart["living_area"]) &
                                    (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"])
                                ]

                                if suitable_aparts.empty:
                                    cannot_offer_to_insert.append((old_apart_id,))
                                else:
                                    suitable_apart = suitable_aparts.iloc[0]
                                    new_apart_id = int(suitable_apart["new_apart_id"])
                                    df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]
                                    offers_to_insert.append((old_apart_id,new_apart_id,))
                            else:
                                suitable_apart = suitable_aparts.iloc[0]
                                new_apart_id = int(suitable_apart["new_apart_id"])
                                df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]
                                offers_to_insert.append((old_apart_id, new_apart_id,))
                        else:
                            cannot_offer_to_insert.append((old_apart_id,))
                    else:
                        suitable_aparts = df_new_apart[
                            (df_new_apart["room_count"] == old_apart["room_count"])  &
                            (df_new_apart["full_living_area"] >= old_apart["full_living_area"])  & 
                            (df_new_apart["total_living_area"] >= old_apart["total_living_area"])  &
                            (df_new_apart["living_area"] >= old_apart["living_area"])  &
                            (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"])  &
                            (df_new_apart["floor"] >= (old_apart["min_floor"])) & 
                            (df_new_apart["floor"] <= (old_apart["max_floor"]))
                        ]

                        # Проверка наличия подходящих квартир
                        if suitable_aparts.empty:
                            suitable_aparts = df_new_apart[
                                (df_new_apart["room_count"] == old_apart["room_count"]) &
                                (df_new_apart["full_living_area"] >= old_apart["full_living_area"]) & 
                                (df_new_apart["total_living_area"] >= old_apart["total_living_area"]) & 
                                (df_new_apart["living_area"] >= old_apart["living_area"]) & 
                                (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"]) &
                                floor_condition
                            ]
                            if suitable_aparts.empty:
                                suitable_aparts = df_new_apart[
                                    (df_new_apart["room_count"] == old_apart["room_count"]) &
                                    (df_new_apart["full_living_area"] >= old_apart["full_living_area"]) &
                                    (df_new_apart["total_living_area"] >= old_apart["total_living_area"]) &
                                    (df_new_apart["living_area"] >= old_apart["living_area"])&
                                    (df_new_apart["for_special_needs_marker"] == old_apart["is_special_needs_marker"])
                                    ]

                                if suitable_aparts.empty:
                                    cannot_offer_to_insert.append((old_apart_id,))

                                else:
                                    suitable_apart = suitable_aparts.iloc[0]
                                    new_apart_id = int(suitable_apart["new_apart_id"])
                                    df_new_apart = df_new_apart[df_new_apart["new_apart_id"]!= new_apart_id]

                                    offers_to_insert.append((old_apart_id, new_apart_id,))
                            else:
                                suitable_apart = suitable_aparts.iloc[0]
                                new_apart_id = int(suitable_apart["new_apart_id"])
                                df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]
                                offers_to_insert.append((old_apart_id, new_apart_id,))

                        else:
                            suitable_apart = suitable_aparts.iloc[0]
                            new_apart_id = int(suitable_apart["new_apart_id"])
                            df_new_apart = df_new_apart[df_new_apart["new_apart_id"] != new_apart_id]

                            offers_to_insert.append(
                                (old_apart_id, new_apart_id,)
                            )

        # Удаление дубликатов из cannot_offer_to_insert
        cannot_offer_to_insert = list(set(cannot_offer_to_insert))
        print('offers_to_insert - ', len(offers_to_insert))
        print('cannot offer to insert - ', len(cannot_offer_to_insert))
        # --- Обновление базы данных ---
        # Для каждой пары old_apart_id и new_apart_id
        for old_apart_id, new_apart_id in offers_to_insert:
            # Создаём JSON с новым апартаментом
            new_aparts_data = {
                str(new_apart_id): {"status_id": 7}
            }
            new_aparts_json = json.dumps(new_aparts_data, ensure_ascii=False)
            print("INSERT INTO public.offer (affair_id, new_aparts, status_id) VALUES (%s, %s, 7)",
                (old_apart_id, new_aparts_json))
            
            # Вставляем новую запись
            cursor.execute(
                "INSERT INTO public.offer (affair_id, new_aparts, status_id) VALUES (%s, %s, 7)",
                (old_apart_id, new_aparts_json)
            )
            print('INSERTED')

        conn.commit()
        res = {'cannot_offer': len(cannot_offer_to_insert), 'offer':  len(offers_to_insert)}
        return res
    except Exception as e:
        print(f"Error: {e}")
        raise


def df_for_aparts(cursor, old_selected_addresses=None, new_selected_addresses=None):
    """
    Retrieves old and new apartment data based on selected addresses.
    
    Args:
        cursor: Database cursor for executing queries
        old_selected_addresses: List of addresses to filter old apartments
        new_selected_addresses: List of addresses with optional sections/ranges to filter new apartments
        
    Returns:
        tuple: (df_old_apart, df_new_apart) - DataFrames containing old and new apartment data
    """
    
    # Query for old apartments
    family_query = """
        SELECT 
            o.affair_id, 
            o.kpu_number,
            o.district, 
            o.municipal_district, 
            o.room_count, 
            o.full_living_area, 
            o.total_living_area, 
            o.living_area, 
            o.is_special_needs_marker, 
            o.min_floor, 
            o.max_floor, 
            o.buying_date,
            ARRAY_AGG(EXTRACT(YEAR FROM AGE(fm.date_of_birth))) AS ages,
            count(o.affair_id) AS members_amount,
            MAX(EXTRACT(YEAR FROM AGE(fm.date_of_birth::timestamp with time zone))) AS oldest,
            o.is_queue,
            o.queue_square,
            o.rank
        FROM 
            old_apart o 
        LEFT JOIN 
            family_member fm ON o.kpu_number = fm.kpu_number 
        WHERE (o.rsm_status <> 'снято' or rsm_status is NULL) and
            o.affair_id NOT IN (
                SELECT affair_id
                FROM offer
                WHERE status_id <> 2
            ) 
    """
    
    old_apart_query_params = []

    # Additional filters for old apartments
    if old_selected_addresses:
        family_query += " AND o.house_address IN %s"
        old_apart_query_params.append(tuple(old_selected_addresses))
    
    # Add GROUP BY and ORDER clauses
    family_query += """
        GROUP BY 
            o.affair_id, 
            o.kpu_number, 
            o.district, 
            o.municipal_district, 
            o.room_count, 
            o.full_living_area, 
            o.total_living_area, 
            o.living_area, 
            o.is_special_needs_marker, 
            o.min_floor, 
            o.max_floor, 
            o.buying_date, 
            o.is_queue, 
            o.queue_square
        ORDER BY 
            o.room_count ASC, 
            (o.full_living_area + o.living_area + (COUNT(fm.family_member_id) / 3.9)) ASC, 
            MAX(EXTRACT(YEAR FROM AGE(fm.date_of_birth::timestamp with time zone))) DESC,
            o.living_area ASC,  
            o.full_living_area ASC, 
            o.total_living_area ASC;
    """

    # Execute old apartments query
    cursor.execute(family_query, old_apart_query_params)
    old_aparts = cursor.fetchall()
    
    # Query for new apartments
    new_apart_query = """
        SELECT
            na.new_apart_id, 
            na.district, 
            na.municipal_district, 
            na.house_address, 
            na.apart_number, 
            na.floor, 
            na.room_count, 
            na.full_living_area, 
            na.total_living_area, 
            na.living_area, 
            na.for_special_needs_marker,
            na.rank  
            FROM
            public.new_apart na
        WHERE
            NOT EXISTS (
                SELECT 1
                FROM public.offer o,
                    jsonb_each(o.new_aparts::jsonb) AS j(key, value)
                WHERE
                j.key = na.new_apart_id::text
                AND (value->>'status_id')::int != 2
            )
            AND (na.status_id NOT IN (12, 13, 15) OR na.status_id IS NULL)
    """

    new_apart_query_params = []

    # Process addresses with apartment ranges for new apartments
    if new_selected_addresses and len(new_selected_addresses) > 0:
        # Extract data from nested list if needed
        if isinstance(new_selected_addresses, list) and len(new_selected_addresses) > 0 and isinstance(new_selected_addresses[0], list):
            new_selected_addresses = new_selected_addresses[0]
        
        address_conditions = []
        
        for address_data in new_selected_addresses:
            if not isinstance(address_data, dict):
                continue
                
            address = address_data.get('address')
            sections = address_data.get('sections', [])
            
            if not sections:
                address_conditions.append("(na.house_address = %s)")
                new_apart_query_params.append(address)
                continue
                
            section_conditions = []
            
            for section_data in sections:
                if not isinstance(section_data, dict):
                    continue
                    
                section = section_data.get('section')
                range_data = section_data.get('range', {})
                
                if isinstance(range_data, dict):
                    try:
                        min_apart = int(range_data.get('from', 0))
                        max_apart = int(range_data.get('to', 0))
                    except (ValueError, TypeError):
                        continue
                elif isinstance(range_data, list) and len(range_data) >= 2:
                    try:
                        min_apart = int(range_data[0])
                        max_apart = int(range_data[1])
                    except (ValueError, TypeError):
                        continue
                else:
                    continue
                
                section_conditions.append(
                    "(na.house_address = %s AND na.apart_number BETWEEN %s AND %s)"
                )
                new_apart_query_params.extend([address, min_apart, max_apart])
            
            if section_conditions:
                address_conditions.append(f"({' OR '.join(section_conditions)})")
            else:
                address_conditions.append("(na.house_address = %s)")
                new_apart_query_params.append(address)

        if address_conditions:
            new_apart_query += " AND (" + " OR ".join(address_conditions) + ")"

    # Add sorting for new apartments
    new_apart_query += " ORDER BY room_count ASC, (full_living_area + living_area), floor, living_area ASC, full_living_area ASC, total_living_area ASC"

    # Execute new apartments query
    cursor.execute(new_apart_query, new_apart_query_params)
    new_aparts = cursor.fetchall()
    
    # Create DataFrames
    df_old_apart = pd.DataFrame(old_aparts, columns=[
        "affair_id", "kpu_number", "district", "municipal_district", "room_count", "full_living_area",
        "total_living_area", "living_area", "is_special_needs_marker", "min_floor", "max_floor", "buying_date", 
        "ages", "members_amount", "oldest", "is_queue", "queue_square", "rank"
    ])

    df_new_apart = pd.DataFrame(new_aparts, columns=[
        "new_apart_id", "district", "municipal_district", "house_address", "apart_number", "floor",
        "room_count", "full_living_area", "total_living_area", "living_area", "for_special_needs_marker", "rank"
    ])

    return df_old_apart, df_new_apart

def save_rank_view_to_excel(writer, history_id, stage_name=None, new_apart_adr=None, old_apart_adr=None, date=False):
    """Функция для обработки и сохранения представления 'rank'"""
    print('Обработка представления: rank')
    try:
        # Создаем уникальное имя листа с учетом этапа
        sheet_name = f"Ранг_{stage_name}" if stage_name else "Ранг"
        
        # Проверяем, существует ли уже такой лист
        if sheet_name in writer.book.sheetnames:
            return  # Пропускаем создание, если лист уже существует
            
        with get_db_connection() as conn:
            query_old_ranked = """
                SELECT
                    old_apart.affair_id as old_apart_id,
                    old_apart.room_count,
                    old_apart.living_area,
                    old_apart.is_special_needs_marker,
                    old_apart.full_living_area,
                    old_apart.total_living_area,
                    old_apart.district,
                    old_apart.municipal_district,
                    old_apart.house_address,
                    old_apart.rank
                FROM old_apart
                WHERE (rsm_status <> 'снято' or rsm_status is null)
            """

            query_new_ranked = """
                SELECT 
                    new_apart.new_apart_id, 
                    new_apart.room_count, 
                    new_apart.living_area, 
                    new_apart.for_special_needs_marker, 
                    new_apart.full_living_area, 
                    new_apart.total_living_area, 
                    new_apart.district, 
                    new_apart.municipal_district, 
                    new_apart.house_address, 
                    new_apart.rank
                FROM new_apart 
                WHERE rank is not NULL
            """
            
            query_old_ranked += f" AND old_apart.history_id = {history_id}"
            query_new_ranked += f" AND new_apart.history_id = {history_id}"

            if date:
                query_old_ranked += " AND old_apart.created_at = (SELECT MAX(created_at) FROM old_apart)"
                query_new_ranked += " AND new_apart.created_at = (SELECT MAX(created_at) FROM new_apart)"
            
            params = {}
            if old_apart_adr:
                query_old_ranked += " AND old_apart.house_address IN %(old_addresses)s"
                params['old_addresses'] = tuple(old_apart_adr)

            if new_apart_adr:
                query_new_ranked += " AND new_apart.house_address IN %(new_addresses)s"
                params['new_addresses'] = tuple(new_apart_adr)

            df_old_ranked = pd.read_sql(query_old_ranked, conn, params=params)
            df_new_ranked = pd.read_sql(query_new_ranked, conn, params=params)

            max_rank_query = """
                SELECT room_count, MAX(rank) as max_rank
                FROM new_apart
                GROUP BY room_count
            """

            max_rank_df = pd.read_sql(max_rank_query, conn)
            max_rank_by_room_count = max_rank_df.set_index("room_count")["max_rank"].to_dict()

            df_combined = pd.concat(
                [
                    df_old_ranked.assign(status="old"),
                    df_new_ranked.assign(status="new"),
                ],
                ignore_index=True,
            )

            df_combined["Ранг"] = df_combined["rank"].astype(int)

            df = (
                df_combined.groupby(["room_count", "Ранг"])
                .agg(
                    Пот_ть=("old_apart_id", "count"),
                    Ресурс=("new_apart_id", "count"),
                )
                .reset_index()
            )

            df["Баланс"] = df["Ресурс"] - df["Пот_ть"]

            def add_totals(df, max_rank_by_room_count):
                                # Инициализируем итоговые переменные
                                total_potency = 0
                                total_resource = 0
                                total_balance = 0

                                # Список для хранения новых строк
                                new_rows = []
                                previous_row = None
                                start_rank = None

                                for i in range(len(df)):
                                    row = df.iloc[i].to_dict()
                                    current_rank = row["Ранг"]
                                    room_count = row["room_count"]
                                    max_rank = max_rank_by_room_count.get(room_count, 0) + 1

                                    if previous_row is not None:
                                        # Проверяем, является ли previous_row['Ранг'] строкой с диапазоном или целым числом
                                        if (
                                            isinstance(previous_row["Ранг"], str)
                                            and "-" in previous_row["Ранг"]
                                        ):
                                            previous_rank = int(
                                                previous_row["Ранг"].split("-")[-1]
                                            )
                                        else:
                                            previous_rank = previous_row["Ранг"]

                                        # Проверяем, можно ли объединять строки
                                        if (
                                            (previous_row["Ресурс"] == 0)
                                            and current_rank != max_rank
                                            and previous_rank != max_rank
                                        ):
                                            # Объединяем строки, если хотя бы одна из них имеет Пот_ть = 0
                                            previous_row["Пот_ть"] += row["Пот_ть"]
                                            previous_row["Ресурс"] += row["Ресурс"]
                                            previous_row["Баланс"] += row["Баланс"]
                                            # Обновляем диапазон рангов
                                            previous_row["Ранг"] = (
                                                f"{start_rank}-{current_rank}"
                                            )
                                        else:
                                            # Добавляем предыдущую строку в список новых строк, если она не None
                                            if previous_row is not None:
                                                new_rows.append(previous_row)

                                            # Устанавливаем новую строку как предыдущую
                                            previous_row = row
                                            start_rank = (
                                                current_rank  # Начало нового диапазона
                                            )
                                    else:
                                        # Устанавливаем первую строку как предыдущую
                                        previous_row = row
                                        start_rank = current_rank

                                # Не забываем добавить последнюю строку после цикла
                                if previous_row is not None:
                                    new_rows.append(previous_row)

                                # Преобразуем список в DataFrame
                                df_new = pd.DataFrame(new_rows)

                                # Пересчитываем итоговые значения
                                total_potency = df_new["Пот_ть"].sum()
                                total_resource = df_new["Ресурс"].sum()
                                total_balance = df_new["Баланс"].sum()

                                # Создаем строку с итогами
                                totals = pd.DataFrame(
                                    [
                                        {
                                            "Ранг": "Итог",
                                            "Пот_ть": total_potency,
                                            "Ресурс": total_resource,
                                            "Баланс": total_balance,
                                        }
                                    ]
                                )

                                # Добавляем строку с итогами в конец нового DataFrame
                                df_with_totals = pd.concat(
                                    [df_new, totals], ignore_index=True
                                )
                                
                                return df_with_totals

            result_data = []
            for room in df["room_count"].unique():
                room_df = df[df["room_count"] == room].copy()
                grouped_df = add_totals(room_df, max_rank_by_room_count)
                grouped_df["room_count"] = room
                result_data.append(grouped_df)

            df_grouped = pd.concat(result_data, ignore_index=True)

            if not df_grouped.empty:
                ws = writer.book.create_sheet(sheet_name)
                current_row, current_col = 1, 1

                header_font = Font(bold=True)
                header_fill = PatternFill(
                    start_color="FFFF99",
                    end_color="FFFF99",
                    fill_type="solid",
                )
                header_alignment = Alignment(horizontal="center")

                for room in df_grouped["room_count"].unique():
                    room_df = df_grouped[df_grouped["room_count"] == room][
                        ["Ранг", "Пот_ть", "Ресурс", "Баланс"]
                    ]
                    ws.cell(
                        row=current_row, column=current_col
                    ).value = f"{room} комната(ы)"
                    ws.cell(
                        row=current_row, column=current_col
                    ).font = header_font
                    ws.cell(
                        row=current_row, column=current_col
                    ).alignment = header_alignment
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=current_col,
                        end_row=current_row,
                        end_column=current_col + 3,
                    )
                    current_row += 1

                    for idx, col_name in enumerate(room_df.columns):
                        cell = ws.cell(
                            row=current_row, column=current_col + idx
                        )
                        cell.value = col_name
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                    current_row += 1

                    for row in dataframe_to_rows(
                        room_df, index=False, header=False
                    ):
                        for idx, value in enumerate(row):
                            ws.cell(
                                row=current_row, column=current_col + idx
                            ).value = value
                        current_row += 1

                    current_row = 1
                    current_col += len(room_df.columns) + 1
    except Exception as e:
        print(f"Ошибка при обработке представления 'rank': {e}")


def save_other_views_to_excel(writer, history_id, stage_name=None, new_apart_adr=None, old_apart_adr=None):
    """Функция для обработки и сохранения других представлений"""
    try:
        views = ["new_apart_all", "res_of_rec", "where_not"]
        with get_db_connection() as conn:
            for view in views:
                print(f"Обработка представления: {view}")
                try:
                    # Для других представлений добавляем stage_name к имени листа
                    sheet_name = f"{view}_{stage_name}" if stage_name else view
                    
                    # Проверяем, существует ли уже такой лист
                    if sheet_name in writer.book.sheetnames:
                        continue
                        
                    query_params = []
                    query = f"SELECT * FROM public.{view}"

                    query += " WHERE 1=1 "
                    if view == "new_apart_all":
                            query += f" AND history_id = ({history_id})"
                            if new_apart_adr:
                                query += ' AND "Новый_адрес" in %s'
                                query_params.append(tuple(new_apart_adr))
                    elif view == 'where_not':
                        if history_id:
                            query += f' AND history_id = ({history_id})'
                            if old_apart_adr:
                                query += ' AND "Старый_адрес" in %s'
                                query_params.append(tuple(old_apart_adr))
                    else:
                        if history_id:
                            query += f" AND id_истории_для_ресурса = {history_id}"
                        if history_id:
                            query += f' AND id_истории_для_кпу = {history_id}'
                        if old_apart_adr:
                                query += ' AND "Старый_адрес" in %s'
                                query_params.append(tuple(old_apart_adr))
                        if new_apart_adr:
                                query += ' AND "Новый_адрес" in %s'
                                query_params.append(tuple(new_apart_adr))
                            
                    print(f"Выполнение запроса для представления {view}")
                    try:
                        df = pd.read_sql(query, conn, params=query_params if query_params else None)
                        df = df.dropna(how="all")
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except Exception as e:
                        print(
                            f"Ошибка выполнения запроса для представления {view}: {e}"
                        )
                except Exception as e:
                    print(e)
    except Exception as e:
        print(f"Ошибка: {e}")

def save_rank_view_to_excel_from_dfs(writer, df_old_apart, df_new_apart, stage_name=None):
    """Функция для обработки и сохранения представления 'rank' из готовых DataFrame"""
    print('Обработка представления: rank из DataFrame')
    try:
        # Создаем уникальное имя листа с учетом этапа
        sheet_name = f"Ранг_{stage_name}" if stage_name else "Ранг"
        
        # Проверяем, существует ли уже такой лист
        if sheet_name in writer.book.sheetnames:
            return  # Пропускаем создание, если лист уже существует

        # Фильтрация данных
        df_old_ranked = df_old_apart[
            (df_old_apart['rank'].notna())
        ].copy()
        
        df_new_ranked = df_new_apart[
            df_new_apart['rank'].notna()
        ].copy()

        # Проверка на пустые DataFrame перед конкатенацией
        if df_old_ranked.empty and df_new_ranked.empty:
            print("Нет данных для отображения")
            return

        # Объединение данных с явной обработкой пустых DF
        dfs_to_concat = []
        if not df_old_ranked.empty:
            dfs_to_concat.append(df_old_ranked.assign(status="old"))
        if not df_new_ranked.empty:
            dfs_to_concat.append(df_new_ranked.assign(status="new"))

        df_combined = pd.concat(dfs_to_concat, ignore_index=True)

        # Преобразование rank в int с обработкой None
        df_combined["Ранг"] = df_combined["rank"].fillna(0).astype(int)

        # Группировка данных
        df = (
            df_combined.groupby(["room_count", "Ранг"])
            .agg(
                Пот_ть=("affair_id", lambda x: x.notna().sum()),
                Ресурс=("new_apart_id", lambda x: x.notna().sum()),
            )
            .reset_index()
        )

        df["Баланс"] = df["Ресурс"] - df["Пот_ть"]

        # Функция для группировки строк с одинаковым балансом
        def add_totals(df, max_rank_by_room_count):
            new_rows = []
            previous_row = None
            start_rank = None

            for i in range(len(df)):
                row = df.iloc[i].to_dict()
                current_rank = row["Ранг"]
                room_count = row["room_count"]
                max_rank = max_rank_by_room_count.get(room_count, 0) + 1

                if previous_row is not None:
                    if (
                        isinstance(previous_row["Ранг"], str)
                        and "-" in previous_row["Ранг"]
                    ):
                        previous_rank = int(previous_row["Ранг"].split("-")[-1])
                    else:
                        previous_rank = previous_row["Ранг"]

                    if (
                        (previous_row["Ресурс"] == 0)
                        and current_rank != max_rank
                        and previous_rank != max_rank
                    ):
                        previous_row["Пот_ть"] += row["Пот_ть"]
                        previous_row["Ресурс"] += row["Ресурс"]
                        previous_row["Баланс"] += row["Баланс"]
                        previous_row["Ранг"] = f"{start_rank}-{current_rank}"
                    else:
                        if previous_row is not None:
                            new_rows.append(previous_row)
                        previous_row = row
                        start_rank = current_rank
                else:
                    previous_row = row
                    start_rank = current_rank

            if previous_row is not None:
                new_rows.append(previous_row)

            df_new = pd.DataFrame(new_rows)
            total_potency = df_new["Пот_ть"].sum()
            total_resource = df_new["Ресурс"].sum()
            total_balance = df_new["Баланс"].sum()

            totals = pd.DataFrame(
                [
                    {
                        "Ранг": "Итог",
                        "Пот_ть": total_potency,
                        "Ресурс": total_resource,
                        "Баланс": total_balance,
                    }
                ]
            )

            df_with_totals = pd.concat([df_new, totals], ignore_index=True)
            return df_with_totals

        # Расчет максимального ранга для каждой комнатности
        max_rank_by_room_count = (
            df_new_ranked.groupby("room_count")["rank"]
            .max()
            .fillna(0)
            .astype(int)
            .to_dict()
        )

        # Обработка данных по комнатностям
        result_data = []
        for room in df["room_count"].unique():
            room_df = df[df["room_count"] == room].copy()
            grouped_df = add_totals(room_df, max_rank_by_room_count)
            grouped_df["room_count"] = room
            result_data.append(grouped_df)

        df_grouped = pd.concat(result_data, ignore_index=True)

        # Запись в Excel
        if not df_grouped.empty:
            ws = writer.book.create_sheet(sheet_name)
            ws.sheet_state = 'visible'  # Убедимся, что лист видим
            
            current_row, current_col = 1, 1

            # Стили для заголовков
            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="FFFF99",
                end_color="FFFF99",
                fill_type="solid",
            )
            header_alignment = Alignment(horizontal="center")

            for room in df_grouped["room_count"].unique():
                room_df = df_grouped[df_grouped["room_count"] == room][
                    ["Ранг", "Пот_ть", "Ресурс", "Баланс"]
                ]
                
                # Заголовок комнатности
                ws.cell(row=current_row, column=current_col).value = f"{int(room)} комната(ы)"
                ws.cell(row=current_row, column=current_col).font = header_font
                ws.cell(row=current_row, column=current_col).alignment = header_alignment
                ws.merge_cells(
                    start_row=current_row,
                    start_column=current_col,
                    end_row=current_row,
                    end_column=current_col + 3,
                )
                current_row += 1

                # Заголовки столбцов
                for idx, col_name in enumerate(room_df.columns):
                    cell = ws.cell(row=current_row, column=current_col + idx)
                    cell.value = col_name
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                current_row += 1

                # Данные
                for row in dataframe_to_rows(room_df, index=False, header=False):
                    for idx, value in enumerate(row):
                        ws.cell(row=current_row, column=current_col + idx).value = value
                    current_row += 1

                current_row = 1
                current_col += len(room_df.columns) + 1

    except Exception as e:
        print(f"Ошибка при обработке представления 'rank' из DataFrame: {e}")
        raise


def waves(data, cursor, conn):
    # Extract addresses from input data
    old_selected_addresses = []
    new_selected_addresses = []

    # Collect old addresses
    for key in sorted(data.keys()):
        if key.startswith('old_apartment_house_address_'):
            for item in data[key]:
                if 'address' in item:
                    old_selected_addresses.append(item['address'])

    # Collect new addresses with ranges
    for key in sorted(data.keys()):
        if key.startswith('new_apartment_house_address_'):
            for item in data[key]:
                print('item', item)
                if 'address' in item:
                    print('ADDED ------------')
                    address_data = {
                        'address': item['address'],
                        'sections': []
                    }
                    
                    if 'sections' in item:
                        for section in item['sections']:
                            section_data = {
                                'section': section.get('section', ''),
                                'range': section.get('range', {})
                            }
                            address_data['sections'].append(section_data)
                    
                    print('address_data', address_data)
                    if address_data['address']:
                        new_selected_addresses.append(address_data)

    # Print address information (for debugging)
    print("\n=== Old addresses ===")
    for i, addr in enumerate(old_selected_addresses, 1):
        print(f"{i}. {addr}")

    print("\n=== New addresses with ranges ===")
    for i, addr in enumerate(new_selected_addresses, 1):
        print(f"{i}. {addr['address']}")
        for j, section in enumerate(addr['sections'], 1):
            print(f"   Section {j}: {section['section']} - Range: {section['range'].get('from', '?')}-{section['range'].get('to', '?')}")

    print(f"\nTotal: {len(old_selected_addresses)} old and {len(new_selected_addresses)} new addresses")

    if (old_selected_addresses == [] or new_selected_addresses == []):
        return None

    df_old_apart, df_new_apart = df_for_aparts(cursor, old_selected_addresses=old_selected_addresses, new_selected_addresses=new_selected_addresses)

    # Создаем переменные для хранения ID квартир
    old_apart_ids_for_history = df_old_apart['affair_id'].tolist()
    new_apart_ids_for_history = df_new_apart['new_apart_id'].tolist()

    # Создаем комбинированный столбец для старых и новых квартир
    df_old_apart["combined_area"] = (df_old_apart["living_area"] + df_old_apart["full_living_area"])
    df_new_apart["combined_area"] = (df_new_apart["living_area"] + df_new_apart["full_living_area"])

    # Присваиваем ранги старым квартирам
    df_old_apart["rank"] = df_old_apart.groupby("room_count")["combined_area"].rank(method="dense").astype(int)

    # Определяем минимальную площадь среди старых квартир 1 ранга
    min_area_rank_1 = df_old_apart[df_old_apart["rank"] == 1].groupby("room_count")["combined_area"].min().to_dict()

    # Создаем словарь для хранения максимальных рангов по количеству комнат среди старых квартир
    max_rank_by_room_count = df_old_apart.groupby("room_count")["rank"].max().to_dict()

    # Присваиваем ранги новым квартирам на основе рангов старых
    df_new_apart["rank"] = 0  # Инициализация колонки для рангов в новых квартирах
    
    for idx, new_row in df_new_apart.iterrows():
        # Определяем минимальную площадь для 1 ранга
        min_area_1_rank = min_area_rank_1.get(new_row["room_count"], float("inf"))

        # Присваиваем ранг 0, если площадь новой квартиры меньше минимальной площади старой квартиры с 1 рангом
        if new_row["combined_area"] < min_area_1_rank:
            df_new_apart.at[idx, "rank"] = 0
        else:
            # Фильтруем старые квартиры, чтобы найти максимальную старую квартиру, которая покрывается новой
            filtered_old = df_old_apart[
                (df_old_apart["room_count"] == new_row["room_count"]) &
                (df_old_apart["living_area"] <= new_row["living_area"]) &
                (df_old_apart["full_living_area"] <= new_row["full_living_area"]) &
                (df_old_apart["total_living_area"] <= new_row["total_living_area"]) & 
                (df_old_apart["is_special_needs_marker"] == new_row["for_special_needs_marker"])
            ]

            if not filtered_old.empty:
                max_rank = filtered_old["rank"].max()
                df_new_apart.at[idx, "rank"] = max_rank

    # Объединяем данные старых и новых квартир
    df_combined = pd.concat([df_old_apart.assign(status="old"), df_new_apart.assign(status="new")], ignore_index=True)

    # Присваиваем индивидуальные ранги без группировки
    df_combined["rank_group"] = df_combined["rank"].astype(int)

    # Обновляем ранги в базе данных для старых и новых квартир
    old_apart_rank_update = list(zip(df_old_apart["rank"], df_old_apart["affair_id"]))
    new_apart_rank_update = list(zip(df_new_apart["rank"], df_new_apart["new_apart_id"]))

    print('old_apart_rank_update', old_apart_rank_update)
    print('new_apart_rank_update', new_apart_rank_update)

    cursor.executemany(
        """UPDATE public.old_apart
                        SET rank = %s
                        WHERE affair_id = %s""",
        old_apart_rank_update,
    )
    cursor.executemany(
        """UPDATE public.new_apart
                        SET rank = %s
                        WHERE new_apart_id = %s
                    """,
        new_apart_rank_update,
    )

    print("Ranking completed successfully")
    print('old_selected_addresses, new_selected_addresses', old_selected_addresses, new_selected_addresses)
    new_selected_addresses_history = list({x['address'] for x in new_selected_addresses})

    cursor.execute(
        """
            INSERT INTO public.history(
                old_house_addresses, 
                new_house_addresses,
                is_wave
            ) 
            VALUES(%s, %s, true)
            RETURNING history_id
        """,
        (old_selected_addresses, new_selected_addresses_history),
    )
    last_history_id = cursor.fetchone()[0]
    conn.commit()

    # Обновление history_id для всех старых квартир, если они есть
    if old_apart_ids_for_history:
        cursor.execute(
            """
            UPDATE public.old_apart
            SET history_id = %s
            WHERE affair_id IN %s
        """,
            (last_history_id, tuple(old_apart_ids_for_history)),
        )
    conn.commit()

    # Обновление history_id для всех новых квартир, если они есть
    if new_apart_ids_for_history:
        cursor.execute(
            """
            UPDATE public.new_apart
            SET history_id = %s
            WHERE new_apart_id IN %s
        """,
            (last_history_id, tuple(new_apart_ids_for_history)),
        )
    conn.commit()


    # Получаем все ключи, которые начинаются с нужных префиксов
    matching_keys = [key for key in data.keys() 
                    if key.startswith(('old_apartment_house_address_', 'new_apartment_house_address_'))]

    # Если нет подходящих ключей, устанавливаем max_i = 0
    if not matching_keys:
        max_i = 0
    else:
        # Извлекаем номера из ключей и находим максимальный
        max_i = max(int(key.split('_')[-1]) for key in matching_keys)

    folders = [Path("uploads")]
    for folder in folders:
        folder.mkdir(parents=True, exist_ok=True)

    output_path = os.path.join(os.getcwd(), "././uploads", f"matching_result_{last_history_id}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        new_addresses_all = [x['address'] for x in new_selected_addresses]
        save_rank_view_to_excel(
            writer=writer,
            history_id=last_history_id,
            stage_name='Общий',
            new_apart_adr=new_addresses_all,
            old_apart_adr=old_selected_addresses
        )
        # Итерируем по всем возможным индексам
        for i in range(1, max_i + 1):
            old_key = f'old_apartment_house_address_{i}'
            new_key = f'new_apartment_house_address_{i}'
            
            print(f"\n=== Итерация {i} ===")
            
            # Получаем старые адреса
            old_addresses = []
            if old_key in data:
                old_addresses = [item['address'] for item in data[old_key] if 'address' in item]
                print(f"Старые адреса ({old_key}): {', '.join(old_addresses)}")
            else:
                print(f"Старые адреса ({old_key}): отсутствуют")
            
            # Получаем новые адреса
            new_addresses = []
            if new_key in data:
                new_addresses = data[new_key]  # Полный объект с секциями и диапазонами
                print(f"Новые адреса ({new_key}):")
                for addr in new_addresses:
                    print(f"  Адрес: {addr.get('address', 'Неизвестный адрес')}")
                    for section in addr.get('sections', []):
                        print(f"    Секция {section.get('section', '?')}: квартиры "
                            f"{section.get('range', {}).get('from', '?')}-{section.get('range', {}).get('to', '?')}")
            else:
                print(f"Новые адреса ({new_key}): отсутствуют")

            if old_addresses and new_addresses:
                df_old_apart_wave, df_new_apart_wave = df_for_aparts(cursor, old_addresses, new_addresses)
                new_apart_adr = [x['address'] for x in new_addresses]
                
                wave_matching(df_new_apart_wave, df_old_apart_wave, cursor, conn)
                
                save_rank_view_to_excel_from_dfs(writer=writer, df_old_apart=df_old_apart_wave, df_new_apart=df_new_apart_wave, stage_name=f"Волна_{i}")
        save_other_views_to_excel(
            writer=writer,
            history_id=last_history_id,
            stage_name='Общий'
        )

    return None


# Example usage
if __name__ == "__main__":    
    # Connect to your database
    conn = get_db_connection()
    cursor = conn.cursor()

    test_data = {
        'old_apartment_house_address_1': [{'address': 'Антонова-Овсеенко ул., д.2 стр. 1'}],
        'new_apartment_house_address_1': [{
            'address': 'Алтуфьевское шоссе, д. 53, корп. 1',
            'sections': []
        }],
        'old_apartment_house_address_2': [{'address': 'Болотниковская ул., д.54 кор.1'}],
        'new_apartment_house_address_2': [{
            'address': 'Алтуфьевское шоссе, д. 53, корп. 1',
            'sections': []
        }]
    }
    
    result = waves(test_data, cursor, conn)
    print(result)
    
    conn.commit()
    conn.close()